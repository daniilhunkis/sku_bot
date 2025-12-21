from datetime import datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def _normalize_cell_value(value):
    # Excel не любит tz-aware datetime, убираем tzinfo
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def build_admin_export(users_rows: Iterable[dict], payments_rows: Iterable[dict]) -> bytes:
    wb = Workbook()
    
    # --- Сегменты пользователей ---
    segments = {
        "Все пользователи": "all",
        "Нажали старт": "started", 
        "Использовали бесплатные": "free_used",
        "Бесплатные закончились": "free_finished",
        "Купили расчёты": "buyers",
        "Неактивные (без расчётов)": "inactive"
    }
    
    for sheet_name, segment in segments.items():
        ws = wb.create_sheet(sheet_name[:31])  # Excel ограничение 31 символ
        
        headers = [
            "ID", "TG User ID", "Создан", "Последний вход",
            "Запусков", "Бесплатных осталось", "Платных осталось",
            "Всего расчётов", "Бесплатных использовано", "Платных использовано"
        ]
        ws.append(headers)
        
        # Стили заголовков
        for cell in ws[1]:
            cell.font = Font(bold=True)
    
    # Основной лист со всеми пользователями
    ws_main = wb.active
    ws_main.title = "Все данные"
    
    main_headers = [
        "tg_user_id", "created_at", "last_seen", "started_count",
        "free_credits", "paid_credits", "free_used", "paid_used", "total_calcs"
    ]
    ws_main.append(main_headers)
    
    for cell in ws_main[1]:
        cell.font = Font(bold=True)
    
    for r in users_rows:
        ws_main.append([_normalize_cell_value(r.get(h)) for h in main_headers])
    
    # --- Платежи ---
    ws_payments = wb.create_sheet("Платежи")
    
    payment_headers = [
        "id", "user_id", "created_at", "amount_rub", 
        "credits", "provider", "external_id", "status"
    ]
    ws_payments.append(payment_headers)
    
    for cell in ws_payments[1]:
        cell.font = Font(bold=True)
    
    for r in payments_rows:
        ws_payments.append([_normalize_cell_value(r.get(h)) for h in payment_headers])
    
    # Автоширина колонок
    for ws in wb.worksheets:
        for column_cells in ws.columns:
            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(10, length + 2)
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
